import pandas as pd
import pyodbc
from openpyxl import load_workbook

def get_user_input():
    server = input("Enter server: ").strip()
    database = input("Enter database: ").strip()
    username = input("Enter username: ").strip()
    password = input("Enter password: ").strip()
    excel_file_path = input("Enter Excel file path: ").strip()

    return server, database, username, password, excel_file_path

def create_table(cursor, table_name, columns):
    create_query = f"CREATE TABLE {table_name} (" + ", ".join(
        [f"[{col}] NVARCHAR(MAX)" for col in columns]) + ")"
    cursor.execute(create_query)

def insert_data(cursor, table_name, df):
    insert_query = f"INSERT INTO {table_name} (" + ", ".join(
        [f"[{col}]" for col in df.columns]) + f") VALUES (" + ", ".join(
        [f"?" for _ in df.columns]) + ")"
    cursor.fast_executemany = True
    cursor.executemany(insert_query, df.fillna('').astype(str).values.tolist())

server, database, username, password, excel_file_path = get_user_input()

if server and database and username and password and excel_file_path:
    try:
        # Connect to Azure SQL Database
        driver = '{ODBC Driver 18 for SQL Server}'
        cnxn = pyodbc.connect(f'DRIVER={driver};SERVER={server};DATABASE={database};UID={username};PWD={password}')
        cursor = cnxn.cursor()

        # Load Excel workbook
        workbook = load_workbook(excel_file_path)
        sheet_names = workbook.sheetnames

        # Import each worksheet into Azure SQL
        for sheet_name in sheet_names:
            # Read worksheet into a Pandas DataFrame
            print(f"Trying to create table '{sheet_name}'")
            df = pd.read_excel(excel_file_path, sheet_name=sheet_name, engine='openpyxl')

            # Replace spaces in column names with underscores
            df.columns = [col.replace(' ', '_') for col in df.columns]

            # Create table in Azure SQL based on DataFrame
            table_name = f"{sheet_name.replace(' ', '_')}_data"
            try:
                create_table(cursor, table_name, df.columns)
                cnxn.commit()
            except pyodbc.Error as e:
                print(f"Error creating table '{table_name}': {e}")
                continue

            # Insert DataFrame data into the table
            try:
                insert_data(cursor, table_name, df)
                cnxn.commit()
                print(f"Inserted data to table '{sheet_name}'")
            except pyodbc.Error as e:
                print(f"Error inserting data into '{table_name}': {e}")
                continue

    except pyodbc.Error as e:
        print(f"Error connecting to Azure SQL Database: {e}")

    finally:
        # Close database connection
        cursor.close()
        cnxn.close()
else:
    print("Please provide all the required details.")
